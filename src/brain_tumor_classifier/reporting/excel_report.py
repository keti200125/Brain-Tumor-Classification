"""Generate an Excel workbook report from global experiment history."""

from __future__ import annotations

from pathlib import Path

import pandas as pd


REQUIRED_HISTORY_COLUMNS = [
    "created_at",
    "experiment_name",
    "trace_name",
    "run_id",
    "config_file",
    "config_hash",
    "report_name",
    "hypothesis",
    "description",
    "tags",
    "model_name",
    "architecture",
    "pretrained_requested",
    "pretrained_used",
    "seed",
    "device",
    "image_size",
    "epochs",
    "batch_size",
    "learning_rate",
    "optimizer",
    "num_workers",
    "validation_fraction",
    "max_train_samples",
    "max_val_samples",
    "max_test_samples",
    "train_size",
    "val_size",
    "test_size",
    "best_epoch",
    "best_val_loss",
    "best_val_accuracy",
    "best_val_f1_weighted",
    "best_val_precision_weighted",
    "best_val_recall_weighted",
    "test_loss",
    "test_accuracy",
    "test_f1_weighted",
    "test_precision_weighted",
    "test_recall_weighted",
    "reloaded_test_accuracy",
    "reloaded_test_f1_weighted",
    "reload_f1_delta",
    "checkpoint_path",
    "last_checkpoint_path",
    "run_dir",
    "plots_dir",
    "metrics_json_path",
    "history_csv_path",
    "status",
    "error_message",
]

EXPERIMENTS_COLUMNS = [
    "created_at",
    "status",
    "experiment_name",
    "trace_name",
    "run_id",
    "model_name",
    "architecture",
    "pretrained_used",
    "image_size",
    "epochs",
    "batch_size",
    "learning_rate",
    "optimizer",
    "train_size",
    "val_size",
    "test_size",
    "best_epoch",
    "best_val_f1_weighted",
    "test_accuracy",
    "test_f1_weighted",
    "test_precision_weighted",
    "test_recall_weighted",
    "reload_f1_delta",
    "checkpoint_path",
    "config_file",
]

BEST_MODELS_COLUMNS = [
    "model_name",
    "best_trace_name",
    "best_run_id",
    "test_accuracy",
    "test_f1_weighted",
    "test_precision_weighted",
    "test_recall_weighted",
    "checkpoint_path",
    "config_file",
]

METRICS_COMPARISON_COLUMNS = [
    "experiment_name",
    "trace_name",
    "model_name",
    "test_accuracy",
    "test_f1_weighted",
    "test_precision_weighted",
    "test_recall_weighted",
    "best_val_f1_weighted",
    "reload_f1_delta",
]

TRACEABILITY_COLUMNS = [
    "experiment_name",
    "trace_name",
    "run_id",
    "config_file",
    "config_hash",
    "hypothesis",
    "description",
    "tags",
    "run_dir",
    "checkpoint_path",
    "history_csv_path",
    "metrics_json_path",
]

FAILED_RUNS_COLUMNS = [
    "created_at",
    "experiment_name",
    "trace_name",
    "run_id",
    "config_file",
    "status",
    "error_message",
]

METRIC_COLUMNS = {
    "best_val_loss",
    "best_val_accuracy",
    "best_val_f1_weighted",
    "best_val_precision_weighted",
    "best_val_recall_weighted",
    "test_loss",
    "test_accuracy",
    "test_f1_weighted",
    "test_precision_weighted",
    "test_recall_weighted",
    "reloaded_test_accuracy",
    "reloaded_test_f1_weighted",
    "reload_f1_delta",
    "learning_rate",
}

COUNT_COLUMNS = {
    "seed",
    "image_size",
    "epochs",
    "batch_size",
    "num_workers",
    "train_size",
    "val_size",
    "test_size",
    "best_epoch",
    "max_train_samples",
    "max_val_samples",
    "max_test_samples",
}


def generate_excel_report(
    history_path: str | Path,
    output_path: str | Path,
) -> Path:
    """Generate the formatted workbook report from experiment history CSV."""
    history_path = Path(history_path)
    output_path = Path(output_path)
    if not history_path.is_file():
        raise FileNotFoundError(f"Experiment history CSV not found: {history_path}")

    dataframe = pd.read_csv(history_path)
    _validate_history_columns(dataframe)
    dataframe = _normalize_types(dataframe)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _build_summary_sheet(dataframe).to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )
        dataframe.loc[:, EXPERIMENTS_COLUMNS].to_excel(
            writer,
            sheet_name="Experiments",
            index=False,
        )
        _build_best_models_sheet(dataframe).to_excel(
            writer,
            sheet_name="Best Models",
            index=False,
        )
        _build_metrics_comparison_sheet(dataframe).to_excel(
            writer,
            sheet_name="Metrics Comparison",
            index=False,
        )
        dataframe.loc[:, TRACEABILITY_COLUMNS].to_excel(
            writer,
            sheet_name="Config Traceability",
            index=False,
        )
        _build_failed_runs_sheet(dataframe).to_excel(
            writer,
            sheet_name="Failed Runs",
            index=False,
        )

    _format_workbook(output_path)
    return output_path


def _build_summary_sheet(dataframe: pd.DataFrame) -> pd.DataFrame:
    successful_runs = dataframe.loc[dataframe["status"] == "success"].copy()
    best_row = _best_metric_row(successful_runs, "test_f1_weighted")

    summary_rows = [
        {"Metric": "Total experiment runs", "Value": int(len(dataframe))},
        {
            "Metric": "Successful runs",
            "Value": int((dataframe["status"] == "success").sum()),
        },
        {
            "Metric": "Failed runs",
            "Value": int((dataframe["status"] != "success").sum()),
        },
        {
            "Metric": "Number of unique models",
            "Value": int(dataframe["model_name"].dropna().nunique()),
        },
        {
            "Metric": "Best model by weighted F1",
            "Value": "" if best_row is None else best_row["model_name"],
        },
        {
            "Metric": "Best experiment name",
            "Value": "" if best_row is None else best_row["experiment_name"],
        },
        {
            "Metric": "Best trace name",
            "Value": "" if best_row is None else best_row["trace_name"],
        },
        {
            "Metric": "Best run ID",
            "Value": "" if best_row is None else best_row["run_id"],
        },
        {
            "Metric": "Best test accuracy",
            "Value": None if best_row is None else best_row["test_accuracy"],
        },
        {
            "Metric": "Best test weighted F1",
            "Value": None if best_row is None else best_row["test_f1_weighted"],
        },
        {
            "Metric": "Best checkpoint path",
            "Value": "" if best_row is None else best_row["checkpoint_path"],
        },
    ]
    return pd.DataFrame(summary_rows)


def _build_best_models_sheet(dataframe: pd.DataFrame) -> pd.DataFrame:
    successful_runs = dataframe.loc[dataframe["status"] == "success"].copy()
    if successful_runs.empty:
        return pd.DataFrame(columns=BEST_MODELS_COLUMNS)

    best_rows: list[dict[str, object]] = []
    for model_name, group in successful_runs.groupby("model_name", dropna=False):
        best_row = _best_metric_row(group, "test_f1_weighted")
        if best_row is None:
            continue
        best_rows.append(
            {
                "model_name": model_name,
                "best_trace_name": best_row["trace_name"],
                "best_run_id": best_row["run_id"],
                "test_accuracy": best_row["test_accuracy"],
                "test_f1_weighted": best_row["test_f1_weighted"],
                "test_precision_weighted": best_row["test_precision_weighted"],
                "test_recall_weighted": best_row["test_recall_weighted"],
                "checkpoint_path": best_row["checkpoint_path"],
                "config_file": best_row["config_file"],
            }
        )
    result = pd.DataFrame(best_rows, columns=BEST_MODELS_COLUMNS)
    if not result.empty:
        result = result.sort_values(
            by=["test_f1_weighted", "test_accuracy"],
            ascending=[False, False],
        ).reset_index(drop=True)
    return result


def _build_metrics_comparison_sheet(dataframe: pd.DataFrame) -> pd.DataFrame:
    successful_runs = dataframe.loc[dataframe["status"] == "success"].copy()
    if successful_runs.empty:
        return pd.DataFrame(columns=METRICS_COMPARISON_COLUMNS)
    result = successful_runs.loc[:, METRICS_COMPARISON_COLUMNS]
    return result.sort_values(
        by=["test_f1_weighted", "test_accuracy"],
        ascending=[False, False],
    ).reset_index(drop=True)


def _build_failed_runs_sheet(dataframe: pd.DataFrame) -> pd.DataFrame:
    failed_runs = dataframe.loc[dataframe["status"] != "success"].copy()
    if failed_runs.empty:
        return pd.DataFrame(columns=FAILED_RUNS_COLUMNS)
    return failed_runs.loc[:, FAILED_RUNS_COLUMNS].reset_index(drop=True)


def _validate_history_columns(dataframe: pd.DataFrame) -> None:
    missing_columns = [
        column for column in REQUIRED_HISTORY_COLUMNS if column not in dataframe.columns
    ]
    if missing_columns:
        raise ValueError(
            "Experiment history CSV is missing required columns: "
            + ", ".join(missing_columns)
        )


def _normalize_types(dataframe: pd.DataFrame) -> pd.DataFrame:
    normalized = dataframe.copy()
    for column in METRIC_COLUMNS | COUNT_COLUMNS:
        if column in normalized.columns:
            normalized[column] = pd.to_numeric(normalized[column], errors="coerce")
    return normalized


def _best_metric_row(
    dataframe: pd.DataFrame,
    metric_column: str,
) -> pd.Series | None:
    candidates = dataframe.dropna(subset=[metric_column])
    if candidates.empty:
        return None
    best_index = candidates[metric_column].astype(float).idxmax()
    return candidates.loc[best_index]


def _format_workbook(output_path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(output_path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    best_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    header_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        headers = [cell.value for cell in worksheet[1]]
        header_lookup = {header: idx + 1 for idx, header in enumerate(headers) if header}
        _apply_number_formats(worksheet, header_lookup)
        _autosize_columns(worksheet, get_column_letter)

        if worksheet.title == "Experiments":
            _highlight_best_metric_cell(
                worksheet,
                header_lookup,
                metric_column="test_f1_weighted",
                fill=best_fill,
            )
        if worksheet.title == "Metrics Comparison":
            _highlight_best_metric_cell(
                worksheet,
                header_lookup,
                metric_column="test_f1_weighted",
                fill=best_fill,
            )

    workbook.save(output_path)


def _apply_number_formats(worksheet: object, header_lookup: dict[str, int]) -> None:
    metric_format = "0.0000"
    count_format = "#,##0"

    for column_name, column_index in header_lookup.items():
        if column_name in METRIC_COLUMNS:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column_index).number_format = metric_format
        elif column_name in COUNT_COLUMNS:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=column_index).number_format = count_format

    if worksheet.title == "Summary":
        value_column = header_lookup.get("Value")
        metric_column = header_lookup.get("Metric")
        if value_column is None or metric_column is None:
            return
        summary_metric_names = {
            "Best test accuracy",
            "Best test weighted F1",
        }
        integer_metric_names = {
            "Total experiment runs",
            "Successful runs",
            "Failed runs",
            "Number of unique models",
        }
        for row in range(2, worksheet.max_row + 1):
            label = worksheet.cell(row=row, column=metric_column).value
            value_cell = worksheet.cell(row=row, column=value_column)
            if label in summary_metric_names:
                value_cell.number_format = metric_format
            elif label in integer_metric_names:
                value_cell.number_format = count_format


def _autosize_columns(worksheet: object, get_column_letter: object) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _highlight_best_metric_cell(
    worksheet: object,
    header_lookup: dict[str, int],
    *,
    metric_column: str,
    fill: object,
) -> None:
    metric_index = header_lookup.get(metric_column)
    if metric_index is None or worksheet.max_row < 2:
        return

    best_row = None
    best_value = None
    for row in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row, column=metric_index).value
        if value is None:
            continue
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            continue
        if best_value is None or numeric_value > best_value:
            best_value = numeric_value
            best_row = row

    if best_row is not None:
        worksheet.cell(row=best_row, column=metric_index).fill = fill
