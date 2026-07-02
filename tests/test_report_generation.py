from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from brain_tumor_classifier.reporting.excel_report import generate_excel_report


class ReportGenerationTest(unittest.TestCase):
    def test_fake_history_csv_generates_workbook_with_required_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            history_path = root / "outputs" / "experiment_history.csv"
            report_path = root / "outputs" / "reports" / "model_report.xlsx"
            history_path.parent.mkdir(parents=True, exist_ok=True)
            self._write_history_csv(history_path)

            generated = generate_excel_report(history_path, report_path)

            self.assertEqual(generated, report_path)
            self.assertTrue(report_path.is_file())
            workbook = load_workbook(report_path)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Summary",
                    "Experiments",
                    "Best Models",
                    "Metrics Comparison",
                    "Config Traceability",
                    "Failed Runs",
                ],
            )

    def _write_history_csv(self, path: Path) -> None:
        columns = [
            "created_at",
            "experiment_name",
            "trace_name",
            "run_id",
            "config_file",
            "config_hash",
            "report_name",
            "hypothesis",
            "description",
            "tags",
            "model_name",
            "architecture",
            "pretrained_requested",
            "pretrained_used",
            "seed",
            "device",
            "image_size",
            "epochs",
            "batch_size",
            "learning_rate",
            "optimizer",
            "num_workers",
            "validation_fraction",
            "max_train_samples",
            "max_val_samples",
            "max_test_samples",
            "train_size",
            "val_size",
            "test_size",
            "best_epoch",
            "best_val_loss",
            "best_val_accuracy",
            "best_val_f1_weighted",
            "best_val_precision_weighted",
            "best_val_recall_weighted",
            "test_loss",
            "test_accuracy",
            "test_f1_weighted",
            "test_precision_weighted",
            "test_recall_weighted",
            "reloaded_test_accuracy",
            "reloaded_test_f1_weighted",
            "reload_f1_delta",
            "checkpoint_path",
            "last_checkpoint_path",
            "run_dir",
            "plots_dir",
            "eda_class_samples_path",
            "eda_class_distribution_path",
            "eda_image_sizes_path",
            "train_vs_val_loss_path",
            "train_vs_val_f1_path",
            "prediction_grid_path",
            "metrics_json_path",
            "history_csv_path",
            "status",
            "error_message",
        ]
        rows = [
            {
                "created_at": "2026-06-30T14:15:22+00:00",
                "experiment_name": "exp_a",
                "trace_name": "trace_a",
                "run_id": "run_a",
                "config_file": "/configs/a.yaml",
                "config_hash": "aaa",
                "report_name": "report a",
                "hypothesis": "hyp a",
                "description": "desc a",
                "tags": "a,b",
                "model_name": "custom_cnn",
                "architecture": "CustomTumorCNN",
                "pretrained_requested": False,
                "pretrained_used": False,
                "seed": 7,
                "device": "cpu",
                "image_size": 32,
                "epochs": 2,
                "batch_size": 4,
                "learning_rate": 0.001,
                "optimizer": "AdamW",
                "num_workers": 0,
                "validation_fraction": 0.5,
                "max_train_samples": "",
                "max_val_samples": "",
                "max_test_samples": "",
                "train_size": 8,
                "val_size": 4,
                "test_size": 4,
                "best_epoch": 2,
                "best_val_loss": 0.21,
                "best_val_accuracy": 0.88,
                "best_val_f1_weighted": 0.87,
                "best_val_precision_weighted": 0.86,
                "best_val_recall_weighted": 0.88,
                "test_loss": 0.24,
                "test_accuracy": 0.89,
                "test_f1_weighted": 0.88,
                "test_precision_weighted": 0.87,
                "test_recall_weighted": 0.89,
                "reloaded_test_accuracy": 0.89,
                "reloaded_test_f1_weighted": 0.88,
                "reload_f1_delta": 0.0,
                "checkpoint_path": "/ckpt/a.pt",
                "last_checkpoint_path": "/ckpt/a_last.pt",
                "run_dir": "/runs/a",
                "plots_dir": "/runs/a/plots",
                "eda_class_samples_path": "/runs/a/plots/eda_class_samples.png",
                "eda_class_distribution_path": "/runs/a/plots/eda_class_distribution.png",
                "eda_image_sizes_path": "/runs/a/plots/eda_image_sizes.png",
                "train_vs_val_loss_path": "/runs/a/plots/train_vs_val_loss.png",
                "train_vs_val_f1_path": "/runs/a/plots/train_vs_val_f1.png",
                "prediction_grid_path": "/runs/a/plots/prediction_grid.png",
                "metrics_json_path": "/runs/a/metrics.json",
                "history_csv_path": "/runs/a/history.csv",
                "status": "success",
                "error_message": "",
            },
            {
                "created_at": "2026-06-30T15:15:22+00:00",
                "experiment_name": "exp_fail",
                "trace_name": "trace_fail",
                "run_id": "run_fail",
                "config_file": "/configs/fail.yaml",
                "config_hash": "bbb",
                "report_name": "report fail",
                "hypothesis": "hyp fail",
                "description": "desc fail",
                "tags": "fail",
                "model_name": "resnet18",
                "architecture": "ResNet18",
                "pretrained_requested": True,
                "pretrained_used": False,
                "seed": 42,
                "device": "cpu",
                "image_size": 224,
                "epochs": 3,
                "batch_size": 8,
                "learning_rate": 0.001,
                "optimizer": "AdamW",
                "num_workers": 0,
                "validation_fraction": 0.5,
                "max_train_samples": "",
                "max_val_samples": "",
                "max_test_samples": "",
                "train_size": "",
                "val_size": "",
                "test_size": "",
                "best_epoch": "",
                "best_val_loss": "",
                "best_val_accuracy": "",
                "best_val_f1_weighted": "",
                "best_val_precision_weighted": "",
                "best_val_recall_weighted": "",
                "test_loss": "",
                "test_accuracy": "",
                "test_f1_weighted": "",
                "test_precision_weighted": "",
                "test_recall_weighted": "",
                "reloaded_test_accuracy": "",
                "reloaded_test_f1_weighted": "",
                "reload_f1_delta": "",
                "checkpoint_path": "",
                "last_checkpoint_path": "",
                "run_dir": "/runs/fail",
                "plots_dir": "/runs/fail/plots",
                "eda_class_samples_path": "",
                "eda_class_distribution_path": "",
                "eda_image_sizes_path": "",
                "train_vs_val_loss_path": "",
                "train_vs_val_f1_path": "",
                "prediction_grid_path": "",
                "metrics_json_path": "/runs/fail/metrics.json",
                "history_csv_path": "/runs/fail/history.csv",
                "status": "failed",
                "error_message": "boom",
            },
        ]
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)


if __name__ == "__main__":
    unittest.main()
