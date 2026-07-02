from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from brain_tumor_classifier.reporting.excel_report import generate_excel_report


class ExcelReportTest(unittest.TestCase):
    def test_generate_excel_report_creates_required_sheets_and_formatting(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            history_path = root / "outputs" / "experiment_history.csv"
            report_path = root / "outputs" / "reports" / "model_report.xlsx"
            history_path.parent.mkdir(parents=True, exist_ok=True)
            self._write_history_csv(history_path)

            output_path = generate_excel_report(history_path, report_path)

            self.assertEqual(output_path, report_path)
            self.assertTrue(report_path.is_file())

            workbook = load_workbook(report_path)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Summary",
                    "Experiments",
                    "Best Models",
                    "Metrics Comparison",
                    "Config Traceability",
                    "Failed Runs",
                ],
            )

            experiments = workbook["Experiments"]
            self.assertEqual(experiments.freeze_panes, "A2")
            self.assertTrue(experiments.auto_filter.ref)
            self.assertTrue(experiments["A1"].font.bold)

            summary_rows = {
                summary_cell.value: value_cell.value
                for summary_cell, value_cell in workbook["Summary"].iter_rows(
                    min_row=2,
                    max_col=2,
                )
            }
            self.assertEqual(summary_rows["Total experiment runs"], 3)
            self.assertEqual(summary_rows["Successful runs"], 2)
            self.assertEqual(summary_rows["Failed runs"], 1)
            self.assertEqual(summary_rows["Best model by weighted F1"], "resnet18")
            self.assertAlmostEqual(summary_rows["Best test weighted F1"], 0.91, places=6)

            best_models = workbook["Best Models"]
            self.assertEqual(best_models.max_row, 3)
            self.assertEqual(best_models["A2"].value, "resnet18")
            self.assertEqual(best_models["B2"].value, "trace_b")

            failed_runs = workbook["Failed Runs"]
            self.assertEqual(failed_runs.max_row, 2)
            self.assertEqual(failed_runs["B2"].value, "exp_fail")
            self.assertEqual(failed_runs["G2"].value, "out of memory")

            highlighted_fills = {
                experiments["T2"].fill.fgColor.rgb,
                experiments["T3"].fill.fgColor.rgb,
                experiments["T4"].fill.fgColor.rgb,
            }
            self.assertIn("00C6EFCE", highlighted_fills)

    def _write_history_csv(self, path: Path) -> None:
        columns = [
            "created_at",
            "experiment_name",
            "trace_name",
            "run_id",
            "config_file",
            "config_hash",
            "report_name",
            "hypothesis",
            "description",
            "tags",
            "model_name",
            "architecture",
            "pretrained_requested",
            "pretrained_used",
            "seed",
            "device",
            "image_size",
            "epochs",
            "batch_size",
            "learning_rate",
            "optimizer",
            "num_workers",
            "validation_fraction",
            "max_train_samples",
            "max_val_samples",
            "max_test_samples",
            "train_size",
            "val_size",
            "test_size",
            "best_epoch",
            "best_val_loss",
            "best_val_accuracy",
            "best_val_f1_weighted",
            "best_val_precision_weighted",
            "best_val_recall_weighted",
            "test_loss",
            "test_accuracy",
            "test_f1_weighted",
            "test_precision_weighted",
            "test_recall_weighted",
            "reloaded_test_accuracy",
            "reloaded_test_f1_weighted",
            "reload_f1_delta",
            "checkpoint_path",
            "last_checkpoint_path",
            "run_dir",
            "plots_dir",
            "metrics_json_path",
            "history_csv_path",
            "status",
            "error_message",
        ]
        rows = [
            {
                "created_at": "2026-06-30T14:15:22+00:00",
                "experiment_name": "exp_a",
                "trace_name": "trace_a",
                "run_id": "run_a",
                "config_file": "/configs/a.yaml",
                "config_hash": "aaa",
                "report_name": "report a",
                "hypothesis": "hyp a",
                "description": "desc a",
                "tags": "a,b",
                "model_name": "custom_cnn",
                "architecture": "CustomTumorCNN",
                "pretrained_requested": False,
                "pretrained_used": False,
                "seed": 7,
                "device": "cpu",
                "image_size": 32,
                "epochs": 2,
                "batch_size": 4,
                "learning_rate": 0.001,
                "optimizer": "AdamW",
                "num_workers": 0,
                "validation_fraction": 0.5,
                "max_train_samples": "",
                "max_val_samples": "",
                "max_test_samples": "",
                "train_size": 8,
                "val_size": 4,
                "test_size": 4,
                "best_epoch": 2,
                "best_val_loss": 0.21,
                "best_val_accuracy": 0.88,
                "best_val_f1_weighted": 0.87,
                "best_val_precision_weighted": 0.86,
                "best_val_recall_weighted": 0.88,
                "test_loss": 0.24,
                "test_accuracy": 0.89,
                "test_f1_weighted": 0.88,
                "test_precision_weighted": 0.87,
                "test_recall_weighted": 0.89,
                "reloaded_test_accuracy": 0.89,
                "reloaded_test_f1_weighted": 0.88,
                "reload_f1_delta": 0.0,
                "checkpoint_path": "/ckpt/a.pt",
                "last_checkpoint_path": "/ckpt/a_last.pt",
                "run_dir": "/runs/a",
                "plots_dir": "/runs/a/plots",
                "metrics_json_path": "/runs/a/metrics.json",
                "history_csv_path": "/runs/a/history.csv",
                "status": "success",
                "error_message": "",
            },
            {
                "created_at": "2026-06-30T15:15:22+00:00",
                "experiment_name": "exp_b",
                "trace_name": "trace_b",
                "run_id": "run_b",
                "config_file": "/configs/b.yaml",
                "config_hash": "bbb",
                "report_name": "report b",
                "hypothesis": "hyp b",
                "description": "desc b",
                "tags": "b,c",
                "model_name": "resnet18",
                "architecture": "ResNet18",
                "pretrained_requested": True,
                "pretrained_used": True,
                "seed": 11,
                "device": "cpu",
                "image_size": 224,
                "epochs": 3,
                "batch_size": 8,
                "learning_rate": 0.0005,
                "optimizer": "AdamW",
                "num_workers": 0,
                "validation_fraction": 0.5,
                "max_train_samples": "",
                "max_val_samples": "",
                "max_test_samples": "",
                "train_size": 16,
                "val_size": 8,
                "test_size": 8,
                "best_epoch": 3,
                "best_val_loss": 0.18,
                "best_val_accuracy": 0.92,
                "best_val_f1_weighted": 0.9,
                "best_val_precision_weighted": 0.91,
                "best_val_recall_weighted": 0.92,
                "test_loss": 0.19,
                "test_accuracy": 0.93,
                "test_f1_weighted": 0.91,
                "test_precision_weighted": 0.92,
                "test_recall_weighted": 0.93,
                "reloaded_test_accuracy": 0.93,
                "reloaded_test_f1_weighted": 0.91,
                "reload_f1_delta": 0.0,
                "checkpoint_path": "/ckpt/b.pt",
                "last_checkpoint_path": "/ckpt/b_last.pt",
                "run_dir": "/runs/b",
                "plots_dir": "/runs/b/plots",
                "metrics_json_path": "/runs/b/metrics.json",
                "history_csv_path": "/runs/b/history.csv",
                "status": "success",
                "error_message": "",
            },
            {
                "created_at": "2026-06-30T16:15:22+00:00",
                "experiment_name": "exp_fail",
                "trace_name": "trace_fail",
                "run_id": "run_fail",
                "config_file": "/configs/fail.yaml",
                "config_hash": "ccc",
                "report_name": "report fail",
                "hypothesis": "hyp fail",
                "description": "desc fail",
                "tags": "fail",
                "model_name": "vgg11",
                "architecture": "VGG11",
                "pretrained_requested": True,
                "pretrained_used": False,
                "seed": 13,
                "device": "cpu",
                "image_size": 224,
                "epochs": 5,
                "batch_size": 8,
                "learning_rate": 0.0003,
                "optimizer": "AdamW",
                "num_workers": 0,
                "validation_fraction": 0.5,
                "max_train_samples": "",
                "max_val_samples": "",
                "max_test_samples": "",
                "train_size": "",
                "val_size": "",
                "test_size": "",
                "best_epoch": "",
                "best_val_loss": "",
                "best_val_accuracy": "",
                "best_val_f1_weighted": "",
                "best_val_precision_weighted": "",
                "best_val_recall_weighted": "",
                "test_loss": "",
                "test_accuracy": "",
                "test_f1_weighted": "",
                "test_precision_weighted": "",
                "test_recall_weighted": "",
                "reloaded_test_accuracy": "",
                "reloaded_test_f1_weighted": "",
                "reload_f1_delta": "",
                "checkpoint_path": "",
                "last_checkpoint_path": "",
                "run_dir": "/runs/fail",
                "plots_dir": "/runs/fail/plots",
                "metrics_json_path": "/runs/fail/metrics.json",
                "history_csv_path": "/runs/fail/history.csv",
                "status": "failed",
                "error_message": "out of memory",
            },
        ]
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)


if __name__ == "__main__":
    unittest.main()
